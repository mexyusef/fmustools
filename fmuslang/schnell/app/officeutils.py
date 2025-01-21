# excel dan word
# f.xls dan f.msdoc
# pip install openpyxl ezsheets
# pip install python-docx
"""
https://pygsheets.readthedocs.io/en/stable/
https://developers.google.com/sheets/api/
A simple, intuitive python library to access google spreadsheets through the Google Sheets API v4. So for example if you have few csv files which you want to export to google sheets and then plot some graphs based on the data. You can use this library to automate that.
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import docx


def doc(filepath):
    return docx.Document(filepath)


def paras(doc):
    return doc.paragraphs


def paralen(doc):
    return len(paras(doc))


def para(doc, i):
    return paras(doc)[i].text


def createxl():
    book = Workbook()
    return book


def savexl(book, filepath):
    book.save(filepath)


def activesheet(book):
    return book.active


def setcell(sheet, value, cell='A1'):
    sheet[cell] = value


def setcellxy(sheet, value, x=1, y=1):
    sheet.cell(row=x, column=y).value = value


def setcells(sheet, x=1, y=1, *series):
    # series adlh list of list
    for i in range(len(series)):
        for j in range(len(series[i])):
            sheet.cell(row=x, column=y).value = series[i][j]


def demo(filepath_input, filepath_output=None, demo_value='Hello World'):
    if not filepath_output:
        filepath_output = filepath_input
    wb = load_workbook(filepath_input)
    sheet = wb.active

    print('read rows')
    # Read by Rows
    for row in sheet.rows:
        for cell in row:
            print(cell.value, end=' ')
        print()

    print('read columns')
    # Read by Column
    for col in sheet.columns:
        for cell in col:
            print(cell.value, end=' ')
        print()

    print('read column A')
    # Read Specific Column
    for col in sheet['A']:
        print(col.value)

    # Read Specific Cell
    print('read cell A1')
    # method 1
    print(sheet['A1'].value)

    print('read cell (1,1)')
    # method 2
    data = sheet.cell(row=1, column=1).value
    print(data)

    print('write and save')
    # Writing and Saving Excel
    sheet.cell(row=1, column=1).value = demo_value
    wb.save(filepath_output)

